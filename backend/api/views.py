from django.shortcuts import render
from rest_framework import viewsets, status
from .models import Form, Submission, Project, Language
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.authtoken.models import Token
from rest_framework.response import Response
from rest_framework.views import APIView
from django.contrib.auth import authenticate
from rest_framework.decorators import action, api_view
import openpyxl
import os
import random
import string
from django.conf import settings
import re

def generate_random_id(length=7):
    return ''.join(random.choices(string.ascii_letters + string.digits, k=length))

def get_ordinal_suffix(n):
    if 10 <= n % 100 <= 20:
        suffix = 'th'
    else:
        suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(n % 10, 'th')
    return f'_{n}{suffix}_choice'

def sanitize_name(name):
    return re.sub(r'[^a-zA-Z0-9_]', '', name)

class FormViewSet(viewsets.ModelViewSet):
    queryset = Form.objects.all()

    def get_serializer_class(self):
        from .serializers import FormSerializer
        return FormSerializer

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        # Check if the request is to update the default language
        update_default_language = request.data.get('update_default_language', False)

        # Update the associated XLSX file
        form = serializer.instance
        questions = request.data.get('questions', form.questions)  # Get questions from the request data or use existing
        default_language = form.default_language
        if default_language:
            default_language_description = f"{default_language.description} ({default_language.subtag})"
        else:
            default_language_description = 'English (en)'

        output_dir = os.path.join(settings.MEDIA_ROOT, 'update')
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f'{form.name}.xlsx')

        wb = openpyxl.Workbook()
        survey_ws = wb.active
        survey_ws.title = 'survey'
        settings_ws = wb.create_sheet(title='settings')
        choices_ws = wb.create_sheet(title='choices')

        # Add headers to the survey sheet
        survey_headers = ['type', 'name', f'label::{default_language_description}', 'required', 'appearance', 'parameters', f'hint::{default_language_description}', 'default', 'guidance_hint', 'hxl', 'constraint_message', 'constraint']
        survey_ws.append(survey_headers)

        # Add headers to the choices sheet
        choices_headers = ['list_name', 'name', 'label']
        choices_ws.append(choices_headers)

        # Add questions to the survey sheet
        for question in questions:
            question_type = question.get('type', 'text')  # Default to 'text' if type is not provided
            question_name = question.get('name', '')
            question_label = question.get('label', '')
            question_required = question.get('required', False)
            question_parameters = question.get('parameters', '')
            question_hint = question.get('hint', '')
            question_default = question.get('default', '')
            question_appearance = question.get('appearance', '')
            question_guidance_hint = question.get('guidance_hint', '')
            question_hxl = question.get('hxl', '')
            question_constraint_message = question.get('constraint_message', '')
            question_constraint = question.get('constraint', '')

            if question_type == 'rating':
                # Generate a single list ID for all select_one questions under this rating question
                list_id = generate_random_id()

                # Add begin_group row
                survey_ws.append(['begin_group', question_name, '', '', 'field-list', '', '', '', '', '', '', ''])

                # Add first select_one row with name <user added name>_header
                survey_ws.append([f'select_one {list_id}', f'{question_name}_header', question_label, '', 'label', '', '', '', '', '', question_constraint_message, question_constraint])

                # Add sub-questions
                for sub_question in question.get('subQuestions', []):
                    sub_question_name = sub_question['name']
                    sub_question_label = sub_question['label']
                    sub_question_required = sub_question['required']
                    sub_question_appearance = sub_question['appearance']
                    sub_question_parameters = sub_question.get('parameters', '')
                    sub_question_constraint_message = question_constraint_message

                    # Ensure the sub-question name starts with an underscore if it starts with a number
                    if sub_question_label and sub_question_label[0].isdigit():
                        sub_question_name = f'_{sub_question_name}'

                    # Sanitize the sub-question name
                    sanitized_sub_question_name = sanitize_name(sub_question_name)

                    # Generate the constraint for the sub-question
                    sub_question_index = sub_question['index']
                    sub_question_constraint = ''
                    for i in range(sub_question_index):
                        other_sub_question_name = question["subQuestions"][i]["name"]
                        if question["subQuestions"][i]["label"][0].isdigit():
                            other_sub_question_name = f'_{other_sub_question_name}'
                        # Sanitize the other sub-question name
                        sanitized_other_sub_question_name = sanitize_name(other_sub_question_name)
                        if sub_question_constraint:
                            sub_question_constraint += ' and '
                        sub_question_constraint += f'${{{sanitized_sub_question_name}}} != ${{{sanitized_other_sub_question_name}}}'

                    survey_ws.append([f'select_one {list_id}', sanitized_sub_question_name, sub_question_label, sub_question_required, sub_question_appearance, sub_question_parameters, '', '', '', '', sub_question_constraint_message, sub_question_constraint])

                # Add end_group row
                survey_ws.append(['end_group', '', '', '', '', '', '', '', '', '', '', ''])

                # Add options to the choices sheet
                options = question.get('options', ['Option 1', 'Option 2'])
                for idx, option in enumerate(options):
                    choices_ws.append([list_id, f'option_{idx + 1}', option])
            else:
                if question_type in ['select_one', 'select_multiple']:
                    list_id = generate_random_id()
                    question_type = f'{question_type} {list_id}'
                    options = question.get('options', [])
                    for idx, option in enumerate(options):
                        choices_ws.append([list_id, f'option_{idx + 1}', option])

                survey_ws.append([question_type, question_name, question_label, question_required, question_appearance, question_parameters, question_hint, question_default, question_guidance_hint, question_hxl, question_constraint_message, question_constraint])

        # Add default_language to the settings sheet if updating default language
        if update_default_language:
            settings_ws.append(['default_language'])
            settings_ws.append([default_language_description])

        wb.save(output_path)

        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def create_form(self, request, pk=None):
        project = Project.objects.get(pk=pk)
        form_name = request.data.get('name')
        questions = request.data.get('questions', [])  # Get questions from the request data
        default_language = request.data.get('default_language')
        other_languages = request.data.get('other_languages', [])

        form = Form.objects.create(project=project, name=form_name, questions=questions, default_language_id=default_language)  # Save questions

        if other_languages:
            form.other_languages.set(other_languages)

        # Create a new workbook and add the survey and settings sheets
        wb = openpyxl.Workbook()
        survey_ws = wb.active
        survey_ws.title = 'survey'
        settings_ws = wb.create_sheet(title='settings')
        choices_ws = wb.create_sheet(title='choices')

        # Add headers to the survey sheet
        survey_ws.append(['type', 'name', 'label', 'required', 'appearance', 'parameters', 'hint', 'default', 'guidance_hint', 'hxl', 'constraint_message', 'constraint'])

        # Add headers to the choices sheet
        choices_ws.append(['list_name', 'name', 'label'])

        # Add questions to the survey sheet
        for question in questions:
            question_type = question.get('type', 'text')  # Default to 'text' if type is not provided
            question_name = question.get('name', '')
            question_label = question.get('label', '')
            question_required = question.get('required', False)
            question_parameters = question.get('parameters', '')
            question_hint = question.get('hint', '')
            question_default = question.get('default', '')
            question_appearance = question.get('appearance', '')
            question_guidance_hint = question.get('guidance_hint', '')
            question_hxl = question.get('hxl', '')
            question_constraint_message = question.get('constraint_message', '')
            question_constraint = question.get('constraint', '')

            if question_type == 'rating':
                # Generate a single list ID for all select_one questions under this rating question
                list_id = generate_random_id()

                # Add begin_group row
                survey_ws.append(['begin_group', question_name, '', '', 'field-list', '', '', '', '', '', '', ''])

                # Add first select_one row with name <user added name>_header
                survey_ws.append([f'select_one {list_id}', f'{question_name}_header', question_label, '', 'label', '', '', '', '', '', question_constraint_message, question_constraint])

                # Add sub-questions
                for sub_question in question.get('subQuestions', []):
                    sub_question_name = sub_question['name']
                    sub_question_label = sub_question['label']
                    sub_question_required = sub_question['required']
                    sub_question_appearance = sub_question['appearance']
                    sub_question_parameters = sub_question.get('parameters', '')
                    sub_question_constraint_message = question_constraint_message

                    # Ensure the sub-question name starts with an underscore if it starts with a number
                    if sub_question_label and sub_question_label[0].isdigit():
                        sub_question_name = f'_{sub_question_name}'

                    # Sanitize the sub-question name
                    sanitized_sub_question_name = sanitize_name(sub_question_name)

                    # Generate the constraint for the sub-question
                    sub_question_index = sub_question['index']
                    sub_question_constraint = ''
                    for i in range(sub_question_index):
                        other_sub_question_name = question["subQuestions"][i]["name"]
                        if question["subQuestions"][i]["label"][0].isdigit():
                            other_sub_question_name = f'_{other_sub_question_name}'
                        # Sanitize the other sub-question name
                        sanitized_other_sub_question_name = sanitize_name(other_sub_question_name)
                        if sub_question_constraint:
                            sub_question_constraint += ' and '
                        sub_question_constraint += f'${{{sanitized_sub_question_name}}} != ${{{sanitized_other_sub_question_name}}}'

                    survey_ws.append([f'select_one {list_id}', sanitized_sub_question_name, sub_question_label, sub_question_required, sub_question_appearance, sub_question_parameters, '', '', '', '', sub_question_constraint_message, sub_question_constraint])

                # Add end_group row
                survey_ws.append(['end_group', '', '', '', '', '', '', '', '', '', '', ''])

                # Add options to the choices sheet
                options = question.get('options', ['Option 1', 'Option 2'])
                for idx, option in enumerate(options):
                    choices_ws.append([list_id, f'option_{idx + 1}', option])
            else:
                if question_type in ['select_one', 'select_multiple']:
                    list_id = generate_random_id()
                    question_type = f'{question_type} {list_id}'
                    options = question.get('options', [])
                    for idx, option in enumerate(options):
                        choices_ws.append([list_id, f'option_{idx + 1}', option])

                survey_ws.append([question_type, question_name, question_label, question_required, question_appearance, question_parameters, question_hint, question_default, question_guidance_hint, question_hxl, question_constraint_message, question_constraint])

        # Save the new XLSX file
        output_dir = os.path.join(settings.MEDIA_ROOT, 'update')
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f'{form.name}.xlsx')
        wb.save(output_path)

        return Response({'message': 'Form created and files generated successfully', 'file_url': request.build_absolute_uri(output_path)}, status=status.HTTP_201_CREATED)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        # Delete the associated XLSX file
        file_path = os.path.join(settings.MEDIA_ROOT, 'update', f'{instance.name}.xlsx')
        if os.path.exists(file_path):
            os.remove(file_path)
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)

class SubmissionViewSet(viewsets.ModelViewSet):
    queryset = Submission.objects.all()

    def get_serializer_class(self):
        from .serializers import SubmissionSerializer
        return SubmissionSerializer

class ProjectViewSet(viewsets.ModelViewSet):
    queryset = Project.objects.all()

    def get_serializer_class(self):
        from .serializers import ProjectSerializer
        return ProjectSerializer

    def get_queryset(self):
        return Project.objects.filter(created_by=self.request.user)

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=['get'])
    def forms(self, request, pk=None):
        project = self.get_object()
        forms = Form.objects.filter(project=project)
        serializer = self.get_serializer(forms, many=True)
        return Response(serializer.data)

class CustomAuthToken(APIView):
    permission_classes = [AllowAny]

    def post(self, request, *args, **kwargs):
        username = request.data.get('username')
        password = request.data.get('password')
        user = authenticate(username=username, password=password)
        if user:
            token, created = Token.objects.get_or_create(user=user)
            return Response({'token': token.key})
        return Response({'error': 'Invalid credentials'}, status=status.HTTP_400_BAD_REQUEST)

class RegisterUser(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        from .serializers import UserSerializer
        serializer = UserSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()
            return Response({'message': 'User registered successfully'}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class LanguageViewSet(viewsets.ModelViewSet):
    queryset = Language.objects.all()

    def get_serializer_class(self):
        from .serializers import LanguageSerializer
        return LanguageSerializer
    permission_classes = [IsAuthenticated]  # Ensure only authenticated users can access this view

@api_view(['PUT'])
def update_translations(request, form_id):
    try:
        form = Form.objects.get(id=form_id)
    except Form.DoesNotExist:
        return Response({'error': 'Form not found'}, status=status.HTTP_404_NOT_FOUND)

    translations = request.data.get('translations', {})

    # Get the actual default language description and subtag
    default_language = form.default_language
    if default_language:
        default_language_description = f"{default_language.description} ({default_language.subtag})"
    else:
        default_language_description = 'English (en)'

    # Update the translations in the form's questions
    for question in form.questions:
        if question['label'] in translations:
            if not question.get('translations'):
                question['translations'] = {}
            question['translations'][default_language_description] = translations[question['label']]
        for sub_question in question.get('subQuestions', []):
            if sub_question['label'] in translations:
                if not sub_question.get('translations'):
                    sub_question['translations'] = {}
                sub_question['translations'][default_language_description] = translations[sub_question['label']]

    form.save()

    # Update the XLSX file
    output_dir = os.path.join(settings.MEDIA_ROOT, 'update')
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, f'{form.name}.xlsx')

    # Load the existing workbook if it exists
    if os.path.exists(output_path):
        wb = openpyxl.load_workbook(output_path)
        survey_ws = wb['survey']
        settings_ws = wb['settings']
        choices_ws = wb['choices']
    else:
        wb = openpyxl.Workbook()
        survey_ws = wb.active
        survey_ws.title = 'survey'
        settings_ws = wb.create_sheet(title='settings')
        choices_ws = wb.create_sheet(title='choices')

    # Clear the existing survey and choices sheets
    if survey_ws.max_row > 1:
        survey_ws.delete_rows(2, survey_ws.max_row - 1)
    if choices_ws.max_row > 1:
        choices_ws.delete_rows(2, choices_ws.max_row - 1)

    # Add headers to the survey sheet if not already present
    survey_headers = ['type', 'name', f'label::{default_language_description}', 'required', 'appearance', 'parameters', f'hint::{default_language_description}', 'default', 'guidance_hint', 'hxl', 'constraint_message', 'constraint']
    if survey_ws.max_row == 0:
        survey_ws.append(survey_headers)
    elif survey_ws.max_row > 0 and [cell.value for cell in survey_ws[1]] != survey_headers:
        survey_ws.delete_rows(1)
        survey_ws.append(survey_headers)

    # Add headers to the choices sheet if not already present
    choices_headers = ['list_name', 'name', 'label']
    if choices_ws.max_row == 0:
        choices_ws.append(choices_headers)
    elif choices_ws.max_row > 0 and [cell.value for cell in choices_ws[1]] != choices_headers:
        choices_ws.delete_rows(1)
        choices_ws.append(choices_headers)

    # Add questions to the survey sheet
    for question in form.questions:
        question_type = question.get('type', 'text')  # Default to 'text' if type is not provided
        question_name = question.get('name', '')
        question_label = translations.get(question.get('label', ''), question.get('label', ''))
        question_required = question.get('required', False)
        question_parameters = question.get('parameters', '')
        question_hint = question.get('hint', '')
        question_default = question.get('default', '')
        question_appearance = question.get('appearance', '')
        question_guidance_hint = question.get('guidance_hint', '')
        question_hxl = question.get('hxl', '')
        question_constraint_message = question.get('constraint_message', '')
        question_constraint = question.get('constraint', '')

        if question_type == 'rating':
            # Generate a single list ID for all select_one questions under this rating question
            list_id = generate_random_id()

            # Add begin_group row
            survey_ws.append(['begin_group', question_name, '', '', 'field-list', '', '', '', '', '', '', ''])

            # Add first select_one row with name <user added name>_header
            survey_ws.append([f'select_one {list_id}', f'{question_name}_header', question_label, '', 'label', '', '', '', '', '', question_constraint_message, question_constraint])

            # Add sub-questions
            for sub_question in question.get('subQuestions', []):
                sub_question_name = sub_question['name']
                sub_question_label = translations.get(sub_question['label'], sub_question['label'])
                sub_question_required = sub_question['required']
                sub_question_appearance = sub_question['appearance']
                sub_question_parameters = sub_question.get('parameters', '')
                sub_question_constraint_message = question_constraint_message

                # Ensure the sub-question name starts with an underscore if it starts with a number
                if sub_question_label and sub_question_label[0].isdigit():
                    sub_question_name = f'_{sub_question_name}'

                # Sanitize the sub-question name
                sanitized_sub_question_name = sanitize_name(sub_question_name)

                # Generate the constraint for the sub-question
                sub_question_index = sub_question['index']
                sub_question_constraint = ''
                for i in range(sub_question_index):
                    other_sub_question_name = question["subQuestions"][i]["name"]
                    if question["subQuestions"][i]["label"][0].isdigit():
                        other_sub_question_name = f'_{other_sub_question_name}'
                    # Sanitize the other sub-question name
                    sanitized_other_sub_question_name = sanitize_name(other_sub_question_name)
                    if sub_question_constraint:
                        sub_question_constraint += ' and '
                    sub_question_constraint += f'${{{sanitized_sub_question_name}}} != ${{{sanitized_other_sub_question_name}}}'

                survey_ws.append([f'select_one {list_id}', sanitized_sub_question_name, sub_question_label, sub_question_required, sub_question_appearance, sub_question_parameters, '', '', '', '', sub_question_constraint_message, sub_question_constraint])

            # Add end_group row
            survey_ws.append(['end_group', '', '', '', '', '', '', '', '', '', '', ''])

            # Add options to the choices sheet
            options = question.get('options', ['Option 1', 'Option 2'])
            for idx, option in enumerate(options):
                choices_ws.append([list_id, f'option_{idx + 1}', option])
        else:
            if question_type in ['select_one', 'select_multiple']:
                list_id = generate_random_id()
                question_type = f'{question_type} {list_id}'
                options = question.get('options', [])
                for idx, option in enumerate(options):
                    choices_ws.append([list_id, f'option_{idx + 1}', option])

            survey_ws.append([question_type, question_name, question_label, question_required, question_appearance, question_parameters, question_hint, question_default, question_guidance_hint, question_hxl, question_constraint_message, question_constraint])

    wb.save(output_path)

    return Response({'message': 'Translations updated successfully'}, status=status.HTTP_200_OK)